import time
import csv
import re
import json
import sqlite3
import argparse
import requests
import logging
import random
import pickle
from urllib.parse import urljoin, urlparse
from datetime import datetime, timedelta
from dataclasses import dataclass, asdict
from typing import List, Dict, Optional, Set
import threading
from concurrent.futures import ThreadPoolExecutor
import psutil
import validators

# Third-party imports
import playwright
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from tqdm import tqdm
from rich.console import Console
from rich.table import Table
from rich.progress import Progress, TextColumn, BarColumn, TaskProgressColumn
from rich.live import Live
import openpyxl
from openpyxl.styles import Font, PatternFill

# Initialize rich console
console = Console()

# Enhanced logging setup
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler("scraper.log"),
        logging.StreamHandler()
    ]
)

@dataclass
class BusinessData:
    """Enhanced business data structure"""
    name: str
    address: str
    phone: str
    website: Optional[str]
    emails: List[str]
    social_media: Dict[str, str]
    rating: Optional[float]
    review_count: Optional[int]
    business_hours: Optional[str]
    price_range: Optional[str]
    category: Optional[str]
    coordinates: Optional[Dict[str, float]]
    status: str
    confidence_score: float
    scraped_at: str
    distance_km: Optional[float] = None

class SessionManager:
    """Manage scraping sessions with persistence"""
    
    def __init__(self, session_file="session.pkl"):
        self.session_file = session_file
        self.session_data = self.load_session()
    
    def save_session(self, data):
        """Save current session data"""
        self.session_data.update(data)
        with open(self.session_file, 'wb') as f:
            pickle.dump(self.session_data, f)
    
    def load_session(self):
        """Load previous session data"""
        try:
            with open(self.session_file, 'rb') as f:
                return pickle.load(f)
        except FileNotFoundError:
            return {"completed_urls": set(), "failed_urls": set(), "results": []}
    
    def is_completed(self, url):
        """Check if URL was already processed"""
        return url in self.session_data.get("completed_urls", set())

class ProxyRotator:
    """Handle proxy rotation for large-scale scraping"""
    
    def __init__(self, proxies=None):
        self.proxies = proxies or []
        self.current_index = 0
    
    def get_proxy(self):
        """Get next proxy in rotation"""
        if not self.proxies:
            return None
        proxy = self.proxies[self.current_index]
        self.current_index = (self.current_index + 1) % len(self.proxies)
        return proxy

class EmailValidator:
    """Validate and score email addresses"""
    
    INVALID_PATTERNS = [
        r'.*noreply.*', r'.*no-reply.*', r'.*admin.*', r'.*test.*',
        r'.*example.*', r'.*dummy.*', r'.*webmaster.*'
    ]
    
    EMAIL_CATEGORIES = {
        'info': r'info@.*',
        'contact': r'contact@.*',
        'sales': r'sales@.*',
        'support': r'support@.*',
        'hello': r'hello@.*',
        'general': r'.*@.*'
    }
    
    @classmethod
    def validate_email(cls, email):
        """Validate email format and check against invalid patterns"""
        if not validators.email(email):
            return False
        
        for pattern in cls.INVALID_PATTERNS:
            if re.match(pattern, email.lower()):
                return False
        
        return True
    
    @classmethod
    def categorize_email(cls, email):
        """Categorize email by type"""
        email_lower = email.lower()
        for category, pattern in cls.EMAIL_CATEGORIES.items():
            if re.match(pattern, email_lower):
                return category
        return 'other'
    
    @classmethod
    def score_email_confidence(cls, email, context=""):
        """Score email confidence based on various factors"""
        score = 0.5  # base score
        
        # Higher score for specific email types
        if 'contact' in email.lower() or 'info' in email.lower():
            score += 0.3
        elif 'sales' in email.lower() or 'hello' in email.lower():
            score += 0.2
        
        # Check if found in contact page
        if 'contact' in context.lower():
            score += 0.2
        
        return min(score, 1.0)

class SocialMediaExtractor:
    """Extract social media profiles from websites"""
    
    SOCIAL_PATTERNS = {
        'facebook': r'(?:https?://)?(?:www\.)?facebook\.com/[\w\.-]+',
        'instagram': r'(?:https?://)?(?:www\.)?instagram\.com/[\w\.-]+',
        'twitter': r'(?:https?://)?(?:www\.)?twitter\.com/[\w\.-]+',
        'linkedin': r'(?:https?://)?(?:www\.)?linkedin\.com/[\w\.-/]+',
        'youtube': r'(?:https?://)?(?:www\.)?youtube\.com/[\w\.-/]+',
        'tiktok': r'(?:https?://)?(?:www\.)?tiktok\.com/@[\w\.-]+'
    }
    
    @classmethod
    def extract_social_media(cls, text):
        """Extract social media URLs from text"""
        social_links = {}
        
        for platform, pattern in cls.SOCIAL_PATTERNS.items():
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                # Clean and validate URL
                url = matches[0]
                if not url.startswith('http'):
                    url = 'https://' + url
                social_links[platform] = url
        
        return social_links

class BusinessIntelligenceExtractor:
    """Extract business intelligence data"""
    
    @staticmethod
    def extract_rating(page):
        """Extract business rating from Google Maps"""
        try:
            rating_element = page.query_selector("span.MW4etd")
            if rating_element:
                return float(rating_element.text_content().strip())
        except (PlaywrightTimeoutError, ValueError):
            return None
        return None
    
    @staticmethod
    def extract_review_count(page):
        """Extract review count"""
        try:
            review_element = page.query_selector("span.UY7F9")
            if review_element:
                review_text = review_element.text_content().strip()
                # Extract number from text like "(123)"
                match = re.search(r'\((\d+)\)', review_text)
                return int(match.group(1)) if match else None
        except (PlaywrightTimeoutError, ValueError):
            return None
        return None
    
    @staticmethod
    def extract_business_hours(page):
        """Extract business hours"""
        try:
            hours_button = page.query_selector("button[data-item-id='oh']")
            if hours_button:
                return hours_button.get_attribute("aria-label")
        except PlaywrightTimeoutError:
            return None
        return None
    
    @staticmethod
    def extract_price_range(page):
        """Extract price range indicators"""
        try:
            price_element = page.query_selector("span.mgr77e")
            if price_element:
                return price_element.text_content().strip()
        except PlaywrightTimeoutError:
            return None
        return None
    
    @staticmethod
    def extract_category(page):
        """Extract business category"""
        try:
            category_element = page.query_selector("button[jsaction='pane.rating.category']")
            if category_element:
                return category_element.text_content().strip()
        except PlaywrightTimeoutError:
            return None
        return None

class EnhancedWebsiteCrawler:
    """Enhanced website crawler with intelligent page detection"""
    
    def __init__(self, session_manager, proxy_rotator):
        self.session_manager = session_manager
        self.proxy_rotator = proxy_rotator
        self.failed_websites = []
    
    def get_headers(self):
        """Get randomized headers"""
        user_agents = [
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        ]
        return {"User-Agent": random.choice(user_agents)}
    
    def fetch_with_retry(self, url, retries=3, base_delay=2):
        """Enhanced fetch with proxy support and better error handling"""
        for attempt in range(1, retries + 1):
            try:
                proxies = None
                proxy = self.proxy_rotator.get_proxy()
                if proxy:
                    proxies = {"http": proxy, "https": proxy}
                
                response = requests.get(
                    url, 
                    headers=self.get_headers(), 
                    timeout=15, 
                    proxies=proxies,
                    verify=False
                )
                
                if response.status_code == 200:
                    return response
                elif response.status_code == 429:  # Rate limited
                    time.sleep(base_delay * attempt * 2)
                    continue
                    
            except Exception as e:
                logging.error(f"Attempt {attempt} failed for {url}: {e}")
                if attempt < retries:
                    delay = base_delay * (2 ** (attempt - 1)) + random.uniform(0, 1)
                    time.sleep(delay)
        
        self.failed_websites.append(url)
        return None
    
    def extract_emails_and_social(self, url, max_pages=5):
        """Extract emails and social media from website"""
        if self.session_manager.is_completed(url):
            return [], {}
        
        visited = set()
        emails = set()
        social_media = {}
        
        def score_page_relevance(page_url, content):
            """Score how likely a page is to contain contact info"""
            score = 0
            page_lower = page_url.lower()
            content_lower = content.lower()
            
            # URL-based scoring
            if any(keyword in page_lower for keyword in ['contact', 'about', 'team']):
                score += 3
            if any(keyword in page_lower for keyword in ['support', 'help', 'reach']):
                score += 2
            
            # Content-based scoring
            if 'contact us' in content_lower or 'get in touch' in content_lower:
                score += 2
            if '@' in content:
                score += 1
                
            return score
        
        def crawl_page(page_url, depth=0):
            if depth >= max_pages or page_url in visited:
                return
            
            visited.add(page_url)
            
            response = self.fetch_with_retry(page_url)
            if not response:
                return
            
            try:
                content = response.text
                
                # Extract emails with context
                found_emails = re.findall(
                    r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", 
                    content
                )
                
                validated_emails = []
                for email in found_emails:
                    if EmailValidator.validate_email(email):
                        validated_emails.append(email)
                
                emails.update(validated_emails)
                
                # Extract social media
                page_social = SocialMediaExtractor.extract_social_media(content)
                social_media.update(page_social)
                
                # Find and prioritize relevant links
                domain = urlparse(url).netloc
                links = re.findall(r'href=["\'](.*?)["\']', content)
                
                scored_links = []
                for link in links:
                    abs_link = urljoin(page_url, link)
                    if (urlparse(abs_link).netloc == domain and 
                        abs_link.startswith("http") and 
                        abs_link not in visited):
                        score = score_page_relevance(abs_link, "")
                        scored_links.append((score, abs_link))
                
                # Crawl highest-scored links first
                scored_links.sort(reverse=True)
                for score, link in scored_links[:3]:  # Limit to top 3 links per page
                    if score > 0:
                        crawl_page(link, depth + 1)
                
            except Exception as e:
                logging.error(f"Error crawling {page_url}: {e}")
        
        crawl_page(url)
        self.session_manager.save_session({"completed_urls": {url}})
        
        return list(emails), social_media

class DatabaseManager:
    """Manage SQLite database for results storage"""
    
    def __init__(self, db_path="businesses.db"):
        self.db_path = db_path
        self.init_database()
    
    def init_database(self):
        """Initialize database tables"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS businesses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                address TEXT,
                phone TEXT,
                website TEXT,
                emails TEXT,
                social_media TEXT,
                rating REAL,
                review_count INTEGER,
                business_hours TEXT,
                price_range TEXT,
                category TEXT,
                coordinates TEXT,
                status TEXT,
                confidence_score REAL,
                scraped_at TEXT,
                distance_km REAL
            )
        """)
        
        conn.commit()
        conn.close()
    
    def save_businesses(self, businesses):
        """Save businesses to database"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        for business in businesses:
            business_dict = asdict(business)
            business_dict['emails'] = json.dumps(business_dict['emails'])
            business_dict['social_media'] = json.dumps(business_dict['social_media'])
            business_dict['coordinates'] = json.dumps(business_dict['coordinates']) if business_dict['coordinates'] else None
            
            cursor.execute("""
                INSERT INTO businesses (
                    name, address, phone, website, emails, social_media,
                    rating, review_count, business_hours, price_range, category,
                    coordinates, status, confidence_score, scraped_at, distance_km
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, tuple(business_dict.values()))
        
        conn.commit()
        conn.close()

class EnhancedBusinessScraper:
    """Main scraper class with all enhanced features"""
    
    def __init__(self, args):
        self.args = args
        self.session_manager = SessionManager()
        self.proxy_rotator = ProxyRotator(args.proxies)
        self.website_crawler = EnhancedWebsiteCrawler(self.session_manager, self.proxy_rotator)
        self.db_manager = DatabaseManager() if args.save_db else None
        self.stats = {
            'total_scraped': 0,
            'emails_found': 0,
            'social_found': 0,
            'websites_processed': 0,
            'start_time': time.time()
        }
        self.playwright = None
        self.browser = None
    
    def setup_browser(self):
        """Setup Playwright browser with enhanced options"""
        self.playwright = sync_playwright().start()
        
        browser_args = [
            "--disable-gpu",
            "--disable-blink-features=AutomationControlled",
            "--disable-web-security",
            "--allow-running-insecure-content",
            "--window-size=1920,1080",
        ]
        
        # Add proxy if available
        proxy = self.proxy_rotator.get_proxy()
        proxy_args = {}
        if proxy:
            proxy_parts = proxy.split('://')
            if len(proxy_parts) > 1:
                proxy_url = proxy_parts[1]
                proxy_args = {
                    "proxy": {
                        "server": f"http://{proxy_url}",
                        "username": proxy_parts[0].split(':')[0] if ':' in proxy_parts[0] else None,
                        "password": proxy_parts[0].split(':')[1] if ':' in proxy_parts[0] else None,
                    }
                }
        
        self.browser = self.playwright.chromium.launch(
            headless=self.args.headless,
            args=browser_args,
            **proxy_args
        )
        
        return self.browser
    
    def extract_coordinates(self, page):
        """Extract GPS coordinates from Google Maps"""
        try:
            current_url = page.url
            # Extract coordinates from URL
            match = re.search(r'@(-?\d+\.\d+),(-?\d+\.\d+)', current_url)
            if match:
                return {'lat': float(match.group(1)), 'lng': float(match.group(2))}
        except Exception:
            pass
        return None
    
    def calculate_distance(self, coords1, coords2):
        """Calculate distance between two coordinates using Haversine formula"""
        if not coords1 or not coords2:
            return None
        
        from math import radians, sin, cos, sqrt, atan2
        
        R = 6371  # Earth's radius in km
        
        lat1, lon1 = radians(coords1['lat']), radians(coords1['lng'])
        lat2, lon2 = radians(coords2['lat']), radians(coords2['lng'])
        
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        
        a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
        c = 2 * atan2(sqrt(a), sqrt(1-a))
        
        return R * c
    
    def scrape_google_maps(self):
        """Enhanced Google Maps scraping using Playwright"""
        search_url = f"https://www.google.com/maps/search/{self.args.query}"
        if self.args.location:
            search_url += f" in {self.args.location}"
        
        browser = self.setup_browser()
        context = browser.new_context(
            viewport={"width": 1920, "height": 1080},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        )
        page = context.new_page()
        
        page.goto(search_url)
        
        # Wait for results to load
        page.wait_for_selector(".Nv2PK", timeout=10000)
        
        results = []
        processed_names = set()  # Prevent duplicates
        bad_domains = [
            "wixsite.com", "weebly.com", "wordpress.com", "blogspot.com",
            "squarespace.com", "godaddysites.com", "site123.me"
        ]
        
        with Progress(
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TaskProgressColumn(),
            TextColumn("• {task.fields[stats]}"),
            console=console
        ) as progress:
            
            scrape_task = progress.add_task(
                "Scraping businesses", 
                total=self.args.max_results,
                stats="Starting..."
            )
            
            while len(results) < self.args.max_results:
                try:
                    businesses = page.query_selector_all(".Nv2PK")
                    
                    for i, biz in enumerate(businesses[len(results):]):
                        if len(results) >= self.args.max_results:
                            break
                        
                        try:
                            # Click on the business using Playwright
                            biz.click()
                            page.wait_for_timeout(random.randint(2000, 4000))  # Random delay
                            
                            # Extract basic info
                            name_element = page.query_selector("h1.DUwDvf")
                            name = name_element.text_content().strip() if name_element else "N/A"
                            
                            # Skip if already processed (duplicate)
                            if name in processed_names:
                                continue
                            processed_names.add(name)
                            
                            # Extract all available data
                            website = self.extract_website(page)
                            address = self.extract_address(page)
                            phone = self.extract_phone(page)
                            
                            # Business intelligence data
                            rating = BusinessIntelligenceExtractor.extract_rating(page)
                            review_count = BusinessIntelligenceExtractor.extract_review_count(page)
                            business_hours = BusinessIntelligenceExtractor.extract_business_hours(page)
                            price_range = BusinessIntelligenceExtractor.extract_price_range(page)
                            category = BusinessIntelligenceExtractor.extract_category(page)
                            coordinates = self.extract_coordinates(page)
                            
                            # Process website for emails and social media
                            emails = []
                            social_media = {}
                            status = "No website"
                            confidence_score = 0.0
                            
                            if website:
                                if any(bad in website for bad in bad_domains):
                                    status = "Low-quality website"
                                    confidence_score = 0.2
                                else:
                                    status = "Processing website"
                                    progress.update(scrape_task, stats=f"Processing {name[:20]}...")
                                    
                                    try:
                                        emails, social_media = self.website_crawler.extract_emails_and_social(
                                            website, max_pages=self.args.max_pages
                                        )
                                        
                                        if emails or social_media:
                                            status = "Data extracted"
                                            confidence_score = 0.8
                                        else:
                                            status = "No contact info found"
                                            confidence_score = 0.4
                                            
                                        self.stats['websites_processed'] += 1
                                        if emails:
                                            self.stats['emails_found'] += len(emails)
                                        if social_media:
                                            self.stats['social_found'] += len(social_media)
                                        
                                    except Exception as e:
                                        logging.error(f"Error processing website {website}: {e}")
                                        status = "Website error"
                                        confidence_score = 0.1
                            
                            # Calculate distance if reference coordinates provided
                            distance_km = None
                            if hasattr(self.args, 'reference_coords') and coordinates:
                                distance_km = self.calculate_distance(self.args.reference_coords, coordinates)
                            
                            business = BusinessData(
                                name=name,
                                address=address,
                                phone=phone,
                                website=website,
                                emails=emails,
                                social_media=social_media,
                                rating=rating,
                                review_count=review_count,
                                business_hours=business_hours,
                                price_range=price_range,
                                category=category,
                                coordinates=coordinates,
                                status=status,
                                confidence_score=confidence_score,
                                scraped_at=datetime.now().isoformat(),
                                distance_km=distance_km
                            )
                            
                            results.append(business)
                            self.stats['total_scraped'] += 1
                            
                            # Update progress
                            progress.update(
                                scrape_task, 
                                advance=1,
                                stats=f"Found: {len([r for r in results if r.emails])} with emails"
                            )
                            
                        except Exception as e:
                            logging.error(f"Error scraping business: {e}")
                            continue
                    
                    # Scroll to load more results
                    page.evaluate("""
                        document.querySelector('div.section-scrollbox')?.scrollBy(0,1000)
                    """)
                    page.wait_for_timeout(random.randint(2000, 4000))
                    
                except Exception as e:
                    logging.error(f"Error in main scraping loop: {e}")
                    break
        
        # Clean up Playwright resources
        page.close()
        context.close()
        browser.close()
        self.playwright.stop()
        
        return results
    
    def extract_website(self, page):
        """Extract website URL with multiple selectors"""
        selectors = [
            "a[data-item-id='authority']",
            "a[href*='http']",
            "button[data-item-id*='website']"
        ]
        
        for selector in selectors:
            try:
                element = page.query_selector(selector)
                if element:
                    return element.get_attribute("href")
            except Exception:
                continue
        return None
    
    def extract_address(self, page):
        """Extract address with fallback selectors"""
        selectors = [
            "button[data-item-id='address']",
            ".AYHFM",
            ".Io6YTe"
        ]
        
        for selector in selectors:
            try:
                element = page.query_selector(selector)
                if element:
                    return element.text_content().strip()
            except Exception:
                continue
        return "N/A"
    
    def extract_phone(self, page):
        """Extract phone with fallback selectors"""
        selectors = [
            "button[data-item-id='phone']",
            "span[data-item-id='phone']",
            "a[href^='tel:']"
        ]
        
        for selector in selectors:
            try:
                element = page.query_selector(selector)
                if element:
                    text = element.text_content().strip()
                    if not text:
                        href = element.get_attribute("href")
                        if href and href.startswith('tel:'):
                            return href.replace("tel:", "")
                    return text
            except Exception:
                continue
        return "N/A"
    
    def retry_failed_websites(self, results):
        """Enhanced retry mechanism"""
        if not self.website_crawler.failed_websites:
            return results
        
        console.print("\n🔄 Retrying failed websites...", style="yellow")
        
        with ThreadPoolExecutor(max_workers=3) as executor:
            retry_futures = {}
            
            for site in set(self.website_crawler.failed_websites):
                future = executor.submit(
                    self.website_crawler.extract_emails_and_social, 
                    site, 
                    max_pages=3
                )
                retry_futures[future] = site
            
            for future in tqdm(retry_futures, desc="Retrying sites"):
                site = retry_futures[future]
                try:
                    emails, social_media = future.result(timeout=30)
                    if emails or social_media:
                        # Update matching result
                        for result in results:
                            if result.website == site:
                                result.emails.extend(emails)
                                result.social_media.update(social_media)
                                result.status = "Retry successful"
                                result.confidence_score = min(result.confidence_score + 0.3, 1.0)
                                break
                except Exception as e:
                    logging.error(f"Retry failed for {site}: {e}")
        
        return results
    
    def export_results(self, results):
        """Export results in multiple formats"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # CSV Export
        if self.args.output_csv:
            csv_filename = f"businesses_{timestamp}.csv"
            with open(csv_filename, 'w', newline='', encoding='utf-8') as f:
                if results:
                    writer = csv.DictWriter(f, fieldnames=list(asdict(results[0]).keys()))
                    writer.writeheader()
                    for result in results:
                        row = asdict(result)
                        row['emails'] = ', '.join(row['emails'])
                        row['social_media'] = json.dumps(row['social_media'])
                        row['coordinates'] = json.dumps(row['coordinates']) if row['coordinates'] else ''
                        writer.writerow(row)
            console.print(f"✅ CSV saved to: {csv_filename}", style="green")
            
        # JSON Export
        if self.args.output_json:
            json_filename = f"businesses_{timestamp}.json"
            with open(json_filename, 'w', encoding='utf-8') as f:
                json.dump([asdict(result) for result in results], f, indent=2, ensure_ascii=False)
            console.print(f"✅ JSON saved to: {json_filename}", style="green")
        
        # Excel Export
        if self.args.output_excel:
            excel_filename = f"businesses_{timestamp}.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Business Data"
            
            if results:
                # Headers
                headers = list(asdict(results[0]).keys())
                for col, header in enumerate(headers, 1):
                    cell = sheet.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                # Data
                for row, result in enumerate(results, 2):
                    result_dict = asdict(result)
                    for col, value in enumerate(result_dict.values(), 1):
                        if isinstance(value, (list, dict)):
                            value = json.dumps(value) if value else ""
                        sheet.cell(row=row, column=col, value=str(value) if value else "")
            
            workbook.save(excel_filename)
            console.print(f"✅ Excel saved to: {excel_filename}", style="green")
        
        # Database Export
        if self.db_manager:
            self.db_manager.save_businesses(results)
            console.print("✅ Data saved to database", style="green")
    
    def display_summary(self, results):
        """Display enhanced results summary"""
        runtime = time.time() - self.stats['start_time']
        
        # Calculate additional statistics
        businesses_with_emails = len([r for r in results if r.emails])
        businesses_with_social = len([r for r in results if r.social_media])
        total_emails = sum(len(r.emails) for r in results)
        avg_confidence = sum(r.confidence_score for r in results) / len(results) if results else 0
        
        # Create summary table
        table = Table(title="📊 Scraping Summary")
        table.add_column("Metric", style="cyan")
        table.add_column("Value", style="magenta")
        
        table.add_row("Total Businesses Scraped", str(len(results)))
        table.add_row("Businesses with Emails", f"{businesses_with_emails} ({businesses_with_emails/len(results)*100:.1f}%)")
        table.add_row("Businesses with Social Media", f"{businesses_with_social} ({businesses_with_social/len(results)*100:.1f}%)")
        table.add_row("Total Emails Found", str(total_emails))
        table.add_row("Total Social Media Profiles", str(self.stats['social_found']))
        table.add_row("Average Confidence Score", f"{avg_confidence:.2f}")
        table.add_row("Websites Processed", str(self.stats['websites_processed']))
        table.add_row("Runtime", f"{runtime:.2f} seconds")
        
        console.print(table)
        
        # Display top businesses by confidence
        if results:
            console.print("\n🏆 Top 5 Businesses by Confidence:", style="bold green")
            top_businesses = sorted(results, key=lambda x: x.confidence_score, reverse=True)[:5]
            
            for i, biz in enumerate(top_businesses, 1):
                console.print(f"{i}. {biz.name} (Score: {biz.confidence_score:.2f})")
                if biz.emails:
                    console.print(f"   📧 Emails: {', '.join(biz.emails[:3])}{'...' if len(biz.emails) > 3 else ''}")
                if biz.website:
                    console.print(f"   🌐 Website: {biz.website}")

def main():
    """Main function with enhanced argument parsing"""
    parser = argparse.ArgumentParser(description="Enhanced Google Maps Business Scraper")
    parser.add_argument("query", help="Search query for businesses")
    parser.add_argument("--location", help="Location for search")
    parser.add_argument("--max-results", type=int, default=50, help="Maximum number of results to scrape")
    parser.add_argument("--max-pages", type=int, default=3, help="Maximum pages to crawl per website")
    parser.add_argument("--headless", action="store_true", help="Run browser in headless mode")
    parser.add_argument("--proxies", nargs="+", help="List of proxies to use for rotation")
    parser.add_argument("--output-csv", action="store_true", help="Export results to CSV")
    parser.add_argument("--output-json", action="store_true", help="Export results to JSON")
    parser.add_argument("--output-excel", action="store_true", help="Export results to Excel")
    parser.add_argument("--save-db", action="store_true", help="Save results to SQLite database")
    parser.add_argument("--reference-lat", type=float, help="Reference latitude for distance calculation")
    parser.add_argument("--reference-lng", type=float, help="Reference longitude for distance calculation")
    
    args = parser.parse_args()
    
    # Set reference coordinates if provided
    if args.reference_lat and args.reference_lng:
        args.reference_coords = {'lat': args.reference_lat, 'lng': args.reference_lng}
    
    # Set default export formats if none specified
    if not any([args.output_csv, args.output_json, args.output_excel, args.save_db]):
        args.output_csv = True
    
    # Initialize and run scraper
    scraper = EnhancedBusinessScraper(args)
    
    console.print("🚀 Starting enhanced business scraping...", style="bold blue")
    console.print(f"🔍 Searching for: {args.query}", style="bold")
    if args.location:
        console.print(f"📍 Location: {args.location}", style="bold")
    
    try:
        results = scraper.scrape_google_maps()
        
        # Retry failed websites
        if scraper.website_crawler.failed_websites:
            results = scraper.retry_failed_websites(results)
        
        # Export results
        scraper.export_results(results)
        
        # Display summary
        scraper.display_summary(results)
        
    except KeyboardInterrupt:
        console.print("\n⏹️  Scraping interrupted by user", style="red")
    except Exception as e:
        console.print(f"\n❌ Error during scraping: {e}", style="red")
        logging.exception("Scraping failed with error:")
    finally:
        # Save session data
        if 'results' in locals():
            scraper.session_manager.save_session({
                "completed_urls": set(url for result in results if result.website),
                "results": [asdict(result) for result in results]
            })

if __name__ == "__main__":
    main()